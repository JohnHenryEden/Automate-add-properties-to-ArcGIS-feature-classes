#-*-encoding:utf-8-*-#
import arcpy
import os
import openpyxl
import re
import sys

reload(sys)
sys.setdefaultencoding('utf-8')

# Data paths:
# 请自行修改成新的路径
arcpy.env.workspace = r"F:\Database\HistoricStreets.gdb"    # ArcPy工作空间
input_ = r"F:\Database\HistoricStreets.gdb\Streets"         # 输入的要素数据类
inputField = u"街道名称"                                    # 输入的连接图片字段名称（不要改）
matchField = u"街道名称"                                    # 输入的连接表中的图片连接字段名称（不要改）
pathField = "Picture"                                       # 输入的连接表中的图片路径字段名称（不要改）
picFolder = r"F:\Database\HistoricStreets\Everything"       # 图片文件夹

xlsFileDir = r"F:\Database\HistoricStreets\HZ-06\HZ-06.xlsx"    # Excel表格


wb = openpyxl.load_workbook(xlsFileDir)
ws = wb.active

rowValue = []
tableValue = []
headValue = []

rowNum = 3
columnNum = 1


for columnNum in range(1, 36):
    rowValue.append(ws.cell(row=2, column=columnNum).value)
headValue.append(rowValue)
rowValue = []


# Append values from the Excel doc to a in-memory list.
for rowNum in range(3, 100):
    for columnNum in range(1, 28):
        rowValue.append(ws.cell(row=rowNum, column=columnNum).value)
    tableValue.append(rowValue)
    rowValue = []


s = str(tableValue).replace('u\'', '\'')
h = str(headValue).replace('u\'', '\'')
hv = headValue.pop()
# print(s.decode("unicode-escape"))

def add_field():
    for i in range(hv.__len__()):
        if hv[i] is not None:
            arcpy.AddField_management(input_, hv[i], "TEXT", field_length=255)
            print(hv[i].encode('utf-8'))

def add_data(): # 添加Excel表格中的数据
# Put table value into the feature class:
# What are we going to do?
    edit = arcpy.da.Editor(arcpy.env.workspace)
    edit.startEditing(False, True)
    edit.startOperation()

    try:
        with arcpy.da.UpdateCursor(input_, '*') as updateCursor:
            for row in updateCursor:
                for entries in tableValue:
                    # Use code as key:
                    # if row[3] == entries[0] and row[4] is None:
                    #    for i in range(4, 14):
                    #        row[i] = entries[i-3]
                    #        updateCursor.updateRow(row)
                    # Use name as key:
                    if row[4] == entries[1] and row[3] is None:
                        for i in range(3, 14):
                            row[i] = entries[i-3]
                            updateCursor.updateRow(row)
                print(str(row).decode("unicode-escape"))

    except Exception as e:
        # If an error occurred, print line number and error message

        tb = sys.exc_info()[2]
        print("Line {0}".format(tb.tb_lineno))
        print(e.message.decode("gbk"))

    edit.stopOperation()
    edit.stopEditing(True)


if arcpy.Exists(r"F:\Database\HistoricStreets.gdb\AttachLinkTable"):
    try:
        arcpy.Delete_management(r"F:\Database\HistoricStreets.gdb\AttachLinkTable")
    except():
        arcpy.AddError("Uanble to purge old table")
arcpy.CreateTable_management(r"F:\Database\HistoricStreets.gdb", "AttachLinkTable")

matchTable = r"F:\Database\HistoricStreets.gdb\AttachLinkTable"

def add_pics(matchtable): # 从图片文件夹中以道路名称为主键添加图片作为附件
    # Write the data into the table:
    arcpy.AddField_management(matchtable, matchField, "TEXT", field_length=255)
    arcpy.AddField_management(matchtable, pathField, "TEXT", field_length=255)

    # Regular Expresson to filter Chinese chars away
    pattern1 = re.compile(r"\\\w+-\d+")
    pattern2 = re.compile(r"\\HZ\d+")
    pattern3 = re.compile(r"\\\w+\d+")
    pattern4 = re.compile(r"\\")

    # Get all the pics in the directory
    with arcpy.da.InsertCursor(matchTable, [matchField, pathField]) as cursor:
        for root, dirs, files in os.walk(picFolder):
            for file in files:
                fileGBK = file.decode("gbk")
                rootGBK = root.decode('gbk')
                roadName = unicode(rootGBK[38:])

                matcher1 = re.search(pattern1, roadName)
                matcher2 = re.search(pattern2, roadName)
                matcher3 = re.search(pattern3, roadName)
                matcher4 = re.search(pattern4, roadName)
                if matcher1 or matcher2 or matcher3 or matcher4:
                    if pattern1.findall(roadName) != []:
                        roadName_ = roadName.strip(pattern1.findall(roadName).pop())
                        #print (roadName_)
                        if '\\' in roadName_:
                            res = roadName_.split("\\")[:-1]
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"))
                                cursor.insertRow([str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"), os.path.join(rootGBK, fileGBK)])
                        else:
                            res = roadName_
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res))
                                cursor.insertRow([str(res), os.path.join(rootGBK, fileGBK)])
                    elif pattern2.findall(roadName) != []:
                        roadName_ = roadName.strip(pattern2.findall(roadName).pop())
                        mid = roadName_[3:]
                        #print (mid)
                        if '\\' in mid:
                            res = mid.split("\\")[:-1]
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"))
                                cursor.insertRow([str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"),os.path.join(rootGBK, fileGBK)])
                        else:
                            res = mid
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res))
                                cursor.insertRow([str(res), os.path.join(rootGBK, fileGBK)])
                    elif pattern3.findall(roadName) != []:
                        roadName_ = roadName.strip(pattern3.findall(roadName).pop())
                        #print (roadName_)
                        if '\\' in roadName_:
                            res = roadName_.split("\\")[:-1]
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"))
                                cursor.insertRow([str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"),os.path.join(rootGBK, fileGBK)])

                        else:
                            res = roadName_
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res))
                                cursor.insertRow([str(res), os.path.join(rootGBK, fileGBK)])
                    else:
                        roadName_ = roadName
                        #print (roadName_)
                        if '\\' in roadName_:
                            res = roadName_.split("\\")[1:2]
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"))
                                cursor.insertRow([str(res).strip('\'').strip('[').strip(']').strip('\'').strip('\'').strip('\'').strip('u').strip('\'').decode("unicode-escape"),os.path.join(rootGBK, fileGBK)])

                        else:
                            res = roadName_
                            if file.endswith(".jpg") is True or file.endswith(".png") is True or file.endswith(".JPG") is True or file.endswith(".PNG") is True:
                                print(str(res))
                                cursor.insertRow([str(res), os.path.join(rootGBK, fileGBK)])

    try:
        arcpy.Delete_management("Streets_ATTACH")
        arcpy.DisableAttachments_management(input_)
    except Exception as e:
        tb = sys.exc_info()[2]
        print("Line {0}".format(tb.tb_lineno))
        print(e.message)

    arcpy.EnableAttachments_management(input_)
    arcpy.AddAttachments_management(input_, inputField, matchtable, matchField, pathField, picFolder)

# add_data()
add_pics(matchTable)
